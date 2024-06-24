from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from django.db import IntegrityError, models
from django.http import Http404, JsonResponse
from django.contrib import messages
from .models import *
from django.contrib.auth import authenticate, login, logout
from datetime import datetime
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied, SuspiciousOperation
from django.core.mail import send_mail
from django.utils import timezone
from django.contrib.auth.models import User
from django.template.loader import render_to_string
import random
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.mail import EmailMessage
import mimetypes
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views.decorators.cache import never_cache

# Administrator!

def is_superuser(user):
    return user.is_authenticated and user.is_superuser

def generate_otp():
    generated_otp = ''.join(str(random.randint(0, 9)) for _ in range(6))
    return generated_otp

@never_cache
def admin_login(request):
    if request.user.is_superuser and request.user.is_superuser:
        return redirect(home)
    try:
        if request.method == "POST":
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_superuser:
                    login(request,user)
                    return redirect(home)
                else:
                    messages.warning(request, "Unauthorized user.")
            else:
                messages.error(request,"Username or Password is incorrect")
                return redirect(admin_login)
    except SuspiciousOperation:
        return render(request, 'sign-in.html')
    
    except PermissionDenied:
        return render(request, 'sign-in.html')
    return render(request, 'sign-in.html')

@never_cache
def forgot_admin_password(request):
    if request.method == "POST":
        email = request.POST.get("email")
        if not email:
            messages.error(request, "Please provide an email.")
            return redirect(admin_login)
        try:
            get_object_or_404(User, email = email)
            otp = generate_otp()
            try:
                admin_otp_object = get_object_or_404(AdminOtp, email = email)
                admin_otp_object.otp = otp
                admin_otp_object.save()
            except Http404:
                AdminOtp.objects.create(
                    email = email, 
                    otp = otp
                )            
                
            subject = "IPCS Admin Password Reset."
            message = f"OTP for resetting admin account password is "            
            html_content = render_to_string('admin_email_templates/forgot_password.html', {
                'subject': subject,
                'message': message,
                "email_subject": subject,
                "email_content": message,
                "otp": otp,
                })            
                    

            email = EmailMessage(
                subject=subject,
                body=html_content,
                from_email="ipcshelpyou@gmail.com",
                to=["ipcshelpyou@gmail.com"]
            )
                        

            # Set the content type of the email to 'text/html'
            email.content_subtype = 'html'

            # Send the email
            email.send()
            messages.success(request, 'OTP for resetting your admin account password has been forwarded to your ipcs email address.')

            return redirect(reset_forgotten_admin_password)
        except Http404:
            messages.error(request, "Invalid email address.")
            return redirect(admin_login)
        
@never_cache
def reset_forgotten_admin_password(request):    
    context = {}
    if request.method == "POST":
        otp = request.POST.get("otp")
        new_password = request.POST.get("new_password")
        repeat_password = request.POST.get("repeat_password")
        if otp:
            try:
                admin_otp_object = get_object_or_404(AdminOtp, otp = int(otp))
                if timezone.now() < admin_otp_object.created_time + timedelta(minutes=5):
                    try:
                        user = get_object_or_404(User, email = admin_otp_object.email)                        
                        if user is not None :
                            if new_password == repeat_password:
                                user.set_password(new_password)
                                messages.success(request, "Password updated successfully.")
                                user.save()
                                logout(request)
                                return redirect(admin_login)
                            else:
                                messages.warning(request, "New passwords does not match.")
                                return redirect(reset_forgotten_admin_password)
                        else:
                            messages.error(request, "Invalid User.")
                            return redirect(admin_login)
                    except Http404:
                        messages.error("Invalid user.")
                        return redirect(admin_login)
                else:
                    messages.error(request, "OTP timed out.")
                    return redirect(admin_login)
            except Http404:
                messages.error(request, "Invalid OTP.")
                return redirect(reset_forgotten_admin_password)
        else:
            messages.error(request, 'OTP field cannot be blank.')
            return redirect(reset_forgotten_admin_password)
    return render(request, 'reset_forgotten_admin_password.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def admin_logout(request):
    logout(request)
    return redirect(admin_login)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def home(request):    
    return render(request, 'admin_home.html')    
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def reset_password(request):
    if request.method == "POST":
        current_password = request.POST.get("current_password")
        new_password = request.POST.get("new_password")
        repeat_password = request.POST.get("repeat_password")
        username = request.user.username
        user = authenticate(request, username = username, password = current_password)
        if user is not None:
            if new_password == repeat_password:
                user.set_password(new_password)
                messages.success(request, "Password updated successfully.")
                user.save()
                logout(request)
                return redirect(admin_login)
            else:
                messages.warning(request, "New passwords do not match.")
                return redirect(reset_password)
        else:
            messages.error(request, "Invalid current password.")
            return redirect(reset_password)
    return render(request, 'reset_password.html')

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def add_product(request):
    if request.method == "POST":
        product_name = request.POST.get("product_name")
        image = request.FILES.get("image")
        link = request.POST.get("link")        
        if product_name != "" and not str(product_name).startswith(" "):
            try:
                Product.objects.create(name = product_name, image = image, link = link)
                messages.success(request, "Product created successfully.")          
                return redirect(add_product)
            except IntegrityError:
                messages.error(request, "This product already exists.")
                return redirect(add_product)  

    return render(request, 'products/add_product.html')            

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_products(request):
    products = Product.objects.all().order_by('name')
    context = {
        "products": products,
    }
    return render(request, 'products/list_products.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_product(request, pk):    
    try:
        product = get_object_or_404(Product, pk = pk)            
    except Http404:
        messages.error(request, "Error. Product not found.")
    context = {
        "product": product
    }
    return render(request, 'products/detail_product.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def update_product(request, pk):    
    context = {}
    try:
        product = get_object_or_404(Product, pk=pk)
        if request.method == "POST":
            product_name = request.POST.get("product_name")
            image = request.FILES.get("image")
            link = request.POST.get("link")
            # updaing product object
            if product_name != "" or not str(product_name).startswith(" "):
                product.name = product_name
                product.link = link
                if image:
                    product.image = image                                                        
                product.save()                        
                messages.success(request, "Product updated successfully")
                return redirect(detail_product, pk)
            else:
                messages.warning(request, "Product name cannot be blank or begin with white space.")
                return redirect(update_product, pk)
        context.update({"product": product})
    except Http404:
        messages.error(request, "Error. Product not found.")
    return render(request, 'products/update_product.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_product(request, pk):    
    try:
        product = get_object_or_404(Product, pk = pk)
        product.delete()
        messages.success(request, "Product deleted successfully.")
    except models.ProtectedError as e:
        messages.error(request, "Cannot delete this product as you have one or more applications with product name as this product.")
        # messages.error(request, e)
        return redirect(detail_product, pk)   
    except Http404:
        messages.error(request, "Error. Product not found.")
    return redirect(list_products)    

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def add_client(request):
    if request.method == "POST":
        name = request.POST.get("name")
        image = request.FILES.get("image")
        try:
            Client.objects.create(name = name, image = image)
            messages.success(request, "Client added successfully.")
            return redirect(add_client)
        except IntegrityError:
            messages.warning(request, "This client already exists.")
            return redirect(add_client)
    return render(request, 'client/add_client.html')        

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_clients(request):    
    clients = Client.objects.all().order_by('name')
    context = {
        "clients": clients,
    }
    return render(request, 'client/list_clients.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_client(request, pk):    
    try:
        client = get_object_or_404(Client, pk = pk)            
    except Http404:
        messages.error(request, "Error. Client not found.")
    context = {
        "client": client
    }
    return render(request, 'client/detail_client.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def update_client(request, pk):    
    context = {}
    try:
        client = get_object_or_404(Client, pk=pk)
        if request.method == "POST":
            name = request.POST.get("name")
            image = request.FILES.get("image")            
            # updaing client object
            client.name = name            
            if image:
                client.image = image            
            client.save()                        
            messages.success(request, "Client updated successfully")
            return redirect(detail_client, pk)                
        context.update({"client": client})
    except Http404:
        messages.error(request, "Error. Client not found.")
    return render(request, 'client/update_client.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_client(request, pk):    
    try:
        client = get_object_or_404(Client, pk = pk)
        client.delete()
        messages.success(request, "Client deleted successfully.")
    except Http404:
        messages.error(request, "Error. Client not found.")
    return redirect(list_clients)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def add_technician(request):    
    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        department = request.POST.get("department")
        residential_location = request.POST.get("residential_location")
        photo = request.FILES.get("photo")
        try:
            Technician.objects.create(
                name = name,
                email = email,
                mobile = mobile,
                department = department,
                residential_location = residential_location,
                photo = photo
            )
            messages.success(request, "Created technician successfully.")        
            return redirect(add_technician)
        except IntegrityError:
            messages.warning(request, "Technician with same contact number or email already exists.")
            return redirect(list_technicians)
    return render(request, 'technician/add_technician.html')    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_technicians(request):    
    technicians = Technician.objects.all().order_by("name")
    context = {
        "technicians": technicians,
    }
    return render(request, 'technician/list_technicians.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_technician(request, pk):    
    context = {}            
    try:
        technician = get_object_or_404(Technician, pk = pk)
        context.update({"technician": technician})          
    except Http404:
        messages.error(request, "Error. Technician not found.")
    return render(request, 'technician/detail_technician.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def update_technician(request, pk):    
    context = {}
    try:   
        technician = get_object_or_404(Technician, pk=pk)
        if request.method == "POST":
            name = request.POST.get("name")
            email = request.POST.get("email")
            mobile = request.POST.get("mobile")
            department = request.POST.get("department")
            residential_location = request.POST.get("residential_location")
            photo = request.FILES.get("photo")

            # updating technician object
            technician.name = name
            technician.email = email
            technician.mobile = mobile
            technician.department = department
            technician.residential_location = residential_location            

            # update photo if provided
            if photo:
                technician.photo = photo

            # save the updated technician object
            try:
                technician.save()
            except IntegrityError:
                messages.warning(request, "Technician with same contact number or email already exists.")
                return redirect(list_technicians)


            messages.success(request, "Technician updated successfully.")
            return redirect(detail_technician, pk)
        context.update({"technician" : technician})        
    except Http404:
        messages.error(request, "Error. Technician not found.")
    return render(request, 'technician/update_technician.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_technician(request, pk):    
    try:
        technician = get_object_or_404(Technician, pk = pk)        
        technician.delete()
        messages.success(request, "Technician deleted successfully.")
    except Http404:
        messages.error(request, "Error. Technician not found.")
    return redirect(list_technicians)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_warranty_applications(request):    
    applications = WarrantyApplication.objects.all().order_by("-application_datetime")

    context = {
        "applications": applications
    }
    return render(request, "warranty/list_warranty_applications.html", context)    


@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def warranty_applications_to_excel(request):
    wb = Workbook()
    ws = wb.active

    warranty_applications = WarrantyApplication.objects.all().order_by('application_datetime')

    headers = ["S.No", "Id", "Application Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Billing Name", "Invoice Number", "Serial Number", "Model Number"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, application in enumerate(warranty_applications, start=1):
        row = [index, application.id, application.application_datetime.strftime("%d-%b-%Y %I:%M %p"), application.customer_name, application.contact_number, application.email_id, application.alternative_number, application.product.name, application.billing_name, application.invoice_number, application.serial_number, application.model_number]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=warranty_applications.xlsx'
    wb.save(response)

    return response

# warranty applications
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_warranty_application(request, pk):    
    context = {}
    try:
        application = get_object_or_404(WarrantyApplication, pk = pk)
        context.update({"application": application})
    except Http404:
        messages.error(request, "Error. Warranty application not found.")
        return redirect('warranty_applications')
    return render(request, "warranty/detail_warranty_application.html", context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_warranty_application(request, pk):
    try:
        application = get_object_or_404(WarrantyApplication, pk = pk)
        application.delete()
        messages.success(request, 'Successfully deleted warranty application.')
        return redirect('warranty_applications')
    except Http404:
        messages.error(request, "Error. Warranty application not found.")
        return redirect('warranty_applications')


@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def complete_warranty(request, pk):    
    context = {}
    try:
        warranty_application = get_object_or_404(WarrantyApplication, pk=pk)
        context.update({"warranty_application": warranty_application})
        if request.method == "POST":
            remark = request.POST.get("remark")
            try:
                finished_warranty = FinishedWarranty.objects.create(
                    id = warranty_application.id,
                    application_datetime = warranty_application.application_datetime,                                                  
                    customer_name = warranty_application.customer_name,
                    contact_number = warranty_application.contact_number,
                    email_id = warranty_application.email_id,
                    alternative_number = warranty_application.alternative_number,
                    product = warranty_application.product,
                    billing_name = warranty_application.billing_name,
                    invoice_number = warranty_application.invoice_number,
                    serial_number = warranty_application.serial_number,
                    model_number = warranty_application.model_number,                
                    address = warranty_application.address,
                    product_complain = warranty_application.product_complain,
                    remark = remark,
                )
                try:
                    get_list_or_404(FinishedWarranty, id = pk)
                    warranty_application.delete()
                    messages.success(request, "Warranty finished successfull.")

                    subject = "Warranty Request Finished"
                    message = "Warranty has been completed."            
                    html_content = render_to_string('admin_email_templates/warranty_request_finished.html', {
                        'subject': subject,
                        'message': message,
                        "email_subject": subject,
                        "email_content": message,
                        "id": finished_warranty.id,
                        "product": finished_warranty.product,
                        "serial_number": finished_warranty.serial_number,
                        "model_number": finished_warranty.model_number,                
                        "billing_name" : finished_warranty.billing_name,
                        "invoice_number": finished_warranty.invoice_number,
                        "product_complain": finished_warranty.product_complain,
                        "customer_name": finished_warranty.customer_name,
                        "email": finished_warranty.email_id,
                        "contact_number": finished_warranty.contact_number,
                        "alternative_number": finished_warranty.alternative_number,
                        "address": finished_warranty.address,
                        "remark": finished_warranty.remark
                        })            
                            

                    email = EmailMessage(
                        subject=subject,
                        body=html_content,
                        from_email="ipcshelpyou@gmail.com",
                        to=[str(finished_warranty.email_id)]
                    )                    

                    # Set the content type of the email to 'text/html'
                    email.content_subtype = 'html'

                    # Send the email
                    email.send()

                    return redirect(list_warranty_applications)
                except Http404:
                    messages.error(request, "Error finishing warranty.")

            except IntegrityError:
                messages.warning(request, "Warranty already Finished.")            
            return redirect(detail_warranty_application, pk)
    except Http404:
        messages.error(request, "Invalid accepted warranty.")
        return redirect(list_warranty_applications)    

@user_passes_test(is_superuser, login_url=admin_login)   
@never_cache
def list_finished_warranties(request):    
    finished_warranties = FinishedWarranty.objects.all().order_by("-finished_datetime")
    context = {
        "finished_warranties": finished_warranties,
    }
    return render(request, 'warranty/list_finished_warranties.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def finished_warranties_to_excel(request):
    wb = Workbook()
    ws = wb.active

    finished_warranties = FinishedWarranty.objects.all().order_by('finished_datetime')

    headers = ["S.No", "Id", "Application Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Billing Name", "Invoice Number", "Serial Number", "Model Number"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, warranty in enumerate(finished_warranties, start=1):
        row = [index, warranty.id, warranty.application_datetime.strftime("%d-%b-%Y %I:%M %p"), warranty.customer_name, warranty.contact_number, warranty.email_id, warranty.alternative_number, warranty.product.name, warranty.billing_name, warranty.invoice_number, warranty.serial_number, warranty.model_number]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=finished_warranty_applications.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def detail_finished_warranty(request, pk):    
    context = {}
    try:
        finished_warranty = get_object_or_404(FinishedWarranty, pk = pk)
        remark = finished_warranty.remark
        context.update({"finished_warranty": finished_warranty})
        if request.method == "POST":
            remark = request.POST.get("remark")
            if remark != finished_warranty.remark:
                finished_warranty.remark = remark
                finished_warranty.save()
                messages.success(request, "Remark Updated.")
                
                subject = "Finished Warranty's Remark Updated."
                message = "Remark of your finished warranty application has been updated."            
                html_content = render_to_string('admin_email_templates/email.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": finished_warranty.id,
                    "product": finished_warranty.product,
                    "serial_number": finished_warranty.serial_number,
                    "model_number": finished_warranty.model_number,                
                    "billing_name" : finished_warranty.billing_name,
                    "invoice_number": finished_warranty.invoice_number,
                    "product_complain": finished_warranty.product_complain,
                    "customer_name": finished_warranty.customer_name,
                    "email": finished_warranty.email_id,
                    "contact_number": finished_warranty.contact_number,
                    "alternative_number": finished_warranty.alternative_number,
                    "address": finished_warranty.address,
                    "remark": finished_warranty.remark
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(finished_warranty.email_id)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(detail_finished_warranty, pk)
    except Http404:
        messages.error(request, "Invalid completed warranty.")
        return redirect(list_finished_warranties)
    return render(request, 'warranty/detail_finished_warranty.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_finished_warranty(request, pk):
    try:
        finished_warranty = get_object_or_404(FinishedWarranty, pk=pk)
        try:
            warranty_application = WarrantyApplication.objects.create(
                id = finished_warranty.id,
                customer_name = finished_warranty.customer_name,
                contact_number = finished_warranty.contact_number,
                email_id = finished_warranty.email_id,
                alternative_number = finished_warranty.alternative_number,
                product = finished_warranty.product,
                billing_name = finished_warranty.billing_name,
                invoice_number = finished_warranty.invoice_number,
                serial_number = finished_warranty.serial_number,
                model_number = finished_warranty.model_number,
                address = finished_warranty.address,
                product_complain = finished_warranty.product_complain
                )
            warranty_application.application_datetime = finished_warranty.application_datetime
            warranty_application.save()
            if warranty_application:
                finished_warranty.delete()
                messages.success(request, "Application has been moved back to Warranty Applications.")
                return redirect(list_finished_warranties)
            else:
                messages.error(request, "Error removing finished warranty application.")
                return redirect(detail_finished_warranty,  pk)   
        except IntegrityError:
            messages.error(request, "Error removing finished warranty application.")
            return redirect(detail_finished_warranty,  pk)                
    except Http404:
        messages.error(request, "Invalid finished warranty item.")
        return redirect(list_finished_warranties)
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_finished_warranty(request, pk):
    try:
        application = get_object_or_404(FinishedWarranty, pk = pk)
        application.delete()
        messages.success(request, 'Successfully deleted finished warranty application.')
        return redirect(list_finished_warranties)
    except Http404:
        messages.error(request, "Error. Finished warranty application not found.")
        return redirect(list_finished_warranties)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def reject_warranty(request, pk):
    context = {}
    try: 
        if request.method == "POST":
            reason = request.POST.get("reason")
            warranty_application = get_object_or_404(WarrantyApplication, pk=pk)
            context.update({"warranty_application": warranty_application})
            try:
                rejected_warranty = RejectedWarranty.objects.create(
                    id = warranty_application.id,
                    application_datetime = warranty_application.application_datetime,                    
                    customer_name = warranty_application.customer_name,
                    contact_number = warranty_application.contact_number,
                    email_id = warranty_application.email_id,
                    alternative_number = warranty_application.alternative_number,
                    product = warranty_application.product,
                    billing_name = warranty_application.billing_name,
                    invoice_number = warranty_application.invoice_number,
                    serial_number = warranty_application.serial_number,
                    model_number = warranty_application.model_number,                
                    address = warranty_application.address,
                    product_complain = warranty_application.product_complain,
                    reason = reason,
                )
                try:
                    get_list_or_404(RejectedWarranty, id = pk)
                    warranty_application.delete()
                    messages.success(request, "Warranty rejected successfull.")                    

                    subject = "Warranty Request Rejected."
                    message = "Warranty has been rejected."            
                    html_content = render_to_string('admin_email_templates/warranty_request_rejected.html', {
                        'subject': subject,
                        'message': message,
                        "email_subject": subject,
                        "email_content": message,
                        "id": rejected_warranty.id,
                        "product": rejected_warranty.product,
                        "serial_number": rejected_warranty.serial_number,
                        "model_number": rejected_warranty.model_number,                
                        "billing_name" : rejected_warranty.billing_name,
                        "invoice_number": rejected_warranty.invoice_number,
                        "product_complain": rejected_warranty.product_complain,
                        "customer_name": rejected_warranty.customer_name,
                        "email": rejected_warranty.email_id,
                        "contact_number": rejected_warranty.contact_number,
                        "alternative_number": rejected_warranty.alternative_number,
                        "address": rejected_warranty.address,
                        "reason": rejected_warranty.reason
                        })            
                            

                    email = EmailMessage(
                        subject=subject,
                        body=html_content,
                        from_email="ipcshelpyou@gmail.com",
                        to=[str(rejected_warranty.email_id)]
                    )                    

                    # Set the content type of the email to 'text/html'
                    email.content_subtype = 'html'

                    # Send the email
                    email.send()

                    return redirect(list_warranty_applications)
                except Http404:
                    messages.error(request, "Error rejecting warranty.")
            except IntegrityError:
                messages.warning(request, "Warranty already rejected.")            
            return redirect(detail_warranty_application, pk)
    except Http404:
        messages.error(request, "Invalid accepted warranty.")
        return redirect(list_warranty_applications)    

@user_passes_test(is_superuser, login_url=admin_login)   
@never_cache
def list_rejected_warranties(request):    
    rejected_warranties = RejectedWarranty.objects.all().order_by("-rejected_datetime")
    context = {
        "rejected_warranties": rejected_warranties,
    }
    return render(request, 'warranty/list_rejected_warranties.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def rejected_warranties_to_excel(request):
    wb = Workbook()
    ws = wb.active

    rejected_warranties = RejectedWarranty.objects.all().order_by('rejected_datetime')

    headers = ["S.No", "Id", "Rejection Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Billing Name", "Invoice Number", "Serial Number", "Model Number", "Application Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, warranty in enumerate(rejected_warranties, start=1):
        row = [index, warranty.id, warranty.rejected_datetime.strftime("%d-%b-%Y %I:%M %p"), warranty.customer_name, warranty.contact_number, warranty.email_id, warranty.alternative_number, warranty.product.name, warranty.billing_name, warranty.invoice_number, warranty.serial_number, warranty.model_number, warranty.application_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rejected_warranty_applications.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_rejected_warranty(request, pk):    
    context = {}
    try:
        rejected_warranty = get_object_or_404(RejectedWarranty, pk = pk)
        reason = rejected_warranty.reason
        context.update({"rejected_warranty": rejected_warranty})
        if request.method == "POST":
            reason = request.POST.get("reason")
            if reason != rejected_warranty.reason:
                rejected_warranty.reason = reason
                rejected_warranty.save()
                messages.success(request, "Reason Updated.")
                
                subject = " Warranty Rejected reason updated."
                message = "Rejection reason for rejected warranty application has been updated."
                html_content = render_to_string('admin_email_templates/email.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": rejected_warranty.id,
                    "product": rejected_warranty.product,
                    "serial_number": rejected_warranty.serial_number,
                    "model_number": rejected_warranty.model_number,                
                    "billing_name" : rejected_warranty.billing_name,
                    "invoice_number": rejected_warranty.invoice_number,
                    "product_complain": rejected_warranty.product_complain,
                    "customer_name": rejected_warranty.customer_name,
                    "email": rejected_warranty.email_id,
                    "contact_number": rejected_warranty.contact_number,
                    "alternative_number": rejected_warranty.alternative_number,
                    "address": rejected_warranty.address,
                    "reason": rejected_warranty.reason
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(rejected_warranty.email_id)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(detail_rejected_warranty, pk)

    except Http404:
        messages.error(request, "Invalid completed warranty.")
        return redirect(list_rejected_warranties)
    return render(request, 'warranty/detail_rejected_warranty.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_rejected_warranty(request, pk):
    try:
        rejected_warranty = get_object_or_404(RejectedWarranty, pk=pk)
        try:
            warranty_application = WarrantyApplication.objects.create(
                id = rejected_warranty.id,                
                customer_name = rejected_warranty.customer_name,
                contact_number = rejected_warranty.contact_number,
                email_id = rejected_warranty.email_id,
                alternative_number = rejected_warranty.alternative_number,
                product = rejected_warranty.product,
                billing_name = rejected_warranty.billing_name,
                invoice_number = rejected_warranty.invoice_number,
                serial_number = rejected_warranty.serial_number,
                model_number = rejected_warranty.model_number,
                address = rejected_warranty.address,
                product_complain = rejected_warranty.product_complain
                )
            warranty_application.application_datetime = rejected_warranty.application_datetime
            warranty_application.save()
            if warranty_application:
                rejected_warranty.delete()
                messages.success(request, "Application has been moved back to Warranty Applications.")
                return redirect(list_rejected_warranties)
            else:
                messages.error(request, "Error removing rejected warranty application.")
                return redirect(detail_rejected_warranty,  pk) 
        except IntegrityError:
            messages.error(request, "Error removing rejected warranty application.")
            return redirect(detail_rejected_warranty,  pk)        
    except Http404:
        messages.error(request, "Invalid rejected warranty item.")
        return redirect(list_rejected_warranties)
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_rejected_warranty(request, pk):
    try:
        application = get_object_or_404(RejectedWarranty, pk = pk)
        application.delete()
        messages.success(request, 'Successfully deleted rejected warranty application.')
        return redirect(list_rejected_warranties)
    except Http404:
        messages.error(request, "Error. Rejected warranty application not found.")
        return redirect(list_rejected_warranties)

# service
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_service_requests(request):    
    service_requests = ServiceRequest.objects.all().order_by("-application_datetime")
    context = {
        "service_requests": service_requests,
    }
    return render(request, 'service/list_service_requests.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def service_requests_to_excel(request):
    wb = Workbook()
    ws = wb.active

    service_requests = ServiceRequest.objects.all().order_by('application_datetime')

    headers = ["S.No", "Id", "Application Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, request in enumerate(service_requests, start=1):
        row = [index, request.id, request.application_datetime.strftime("%d-%b-%Y %I:%M %p"), request.customer_name, request.contact_number, request.email, request.alternative_number, request.product.name, request.serial_number, request.model_number]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=service_requests.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_service_request(request, pk):    
    context = {} 
    try:
        service_request = get_object_or_404(ServiceRequest, pk = pk)                
        context.update({"service_request": service_request})
    except Http404:
        messages.error(request, "Error. Service request not found.")
        return redirect(list_service_requests   )
    return render(request, 'service/detail_service_request.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_service_request(request, pk):
    try:
        service_request = get_object_or_404(ServiceRequest, pk = pk)
        service_request.delete()
        messages.success(request, 'Successfully deleted service request.')
        return redirect(list_service_requests)
    except Http404:
        messages.error(request, "Error. Service request not found.")
        return redirect(list_service_requests)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def accept_service(request, pk):    
    # context = {}        
    try:
        get_object_or_404(AcceptedService, pk = pk)
        messages.error(request, "Service already has been already accepted.")
        return redirect(list_accepted_services)
    except:
        pass
    try:
        service_request = get_object_or_404(ServiceRequest, pk = pk)    
        try:
            accepted_service = AcceptedService.objects.create(
                id = service_request.id,
                application_datetime = service_request.application_datetime,
                customer_name = service_request.customer_name,
                address_site = service_request.address_site,
                email = service_request.email,                
                product = service_request.product,
                contact_number = service_request.contact_number,
                alternative_number = service_request.alternative_number,
                serial_number = service_request.serial_number,
                model_number = service_request.model_number,
                service_description = service_request.service_description,
            )
            if accepted_service:
                service_request.delete()                      
                messages.success(request, "Service request has been accepted.")
                
                subject = "Service Request Accepted"
                message = "Your service request has been accepted."
                html_content = render_to_string('admin_email_templates/service_request_accepted.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": accepted_service.id,
                    "product": accepted_service.product,
                    "serial_number": accepted_service.serial_number,
                    "model_number": accepted_service.model_number,                                    
                    "service_description": accepted_service.service_description,
                    "customer_name": accepted_service.customer_name,
                    "email": accepted_service.email,
                    "contact_number": accepted_service.contact_number,
                    "alternative_number": accepted_service.alternative_number,
                    "address": accepted_service.address_site,
                    "remark": accepted_service.remark
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(accepted_service.email)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(list_service_requests)            
            else:
                messages.error(request, "Error accepting service")
                return redirect(detail_service_request)
        except IntegrityError:                
            messages.warning(request, "Service has been aleady accepted.")
            return redirect(detail_accepted_service, pk)                
    except Http404:
        messages.error(request, "Invalid service request.")
        return redirect(list_service_requests) 

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_accepted_services(request):    
    accepted_services = AcceptedService.objects.all().order_by("accepted_datetime")
    context = {"accepted_services" : accepted_services}
    return render(request, 'service/list_accepted_services.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def accepted_services_to_excel(request):
    wb = Workbook()
    ws = wb.active

    accepted_services = AcceptedService.objects.all().order_by('accepted_datetime')

    headers = ["S.No", "Id", "Acception Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number", "Application Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, service in enumerate(accepted_services, start=1):
        row = [index, service.id, service.accepted_datetime.strftime("%d-%b-%Y %I:%M %p"), service.customer_name, service.contact_number, service.email, service.alternative_number, service.product.name, service.serial_number, service.model_number, service.application_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=accepted_services.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_accepted_service(request, pk):
    try:
        accepted_service = get_object_or_404(AcceptedService, pk = pk)
        context = {"accepted_service" : accepted_service}
        if request.method == "POST":
            remark = request.POST.get("remark")
            if remark != accepted_service.remark:
                accepted_service.remark = remark
                accepted_service.save()
                messages.success(request, "Remark Updated.")
                
                subject = "Accepted Service Request's Remark Updated."
                message = "Remark of accepted service request has been updated."
                html_content = render_to_string('admin_email_templates/email.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": accepted_service.id,
                    "product": accepted_service.product,
                    "serial_number": accepted_service.serial_number,
                    "model_number": accepted_service.model_number,                                    
                    "service_description": accepted_service.service_description,
                    "customer_name": accepted_service.customer_name,
                    "email": accepted_service.email,
                    "contact_number": accepted_service.contact_number,
                    "alternative_number": accepted_service.alternative_number,
                    "address": accepted_service.address_site,
                    "remark": accepted_service.remark
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(accepted_service.email)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(detail_accepted_service, pk)

    except Http404:
        messages.error(request, "Invalid accepted service.")
        return redirect(list_accepted_services)
    return render(request, 'service/detail_accepted_service.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_accepted_service(request, pk):    
    try:
        accepted_service = get_object_or_404(AcceptedService, pk=pk)
        try:
            service_request = ServiceRequest.objects.create(
                id = accepted_service.id,
                customer_name = accepted_service.customer_name,
                contact_number = accepted_service.contact_number,
                email = accepted_service.email,            
                alternative_number = accepted_service.alternative_number,
                address_site = accepted_service.address_site,
                product = accepted_service.product,                
                serial_number = accepted_service.serial_number,
                model_number = accepted_service.model_number,
                service_description = accepted_service.service_description,
            )
            service_request.application_datetime = accepted_service.application_datetime
            service_request.save()
            if service_request:
                accepted_service.delete()
                messages.success(request, "Moved the application back to Service Requests.")            
                return redirect(list_accepted_services)
            else:
                messages.error(request, "Error removing the accepted warranty application.")            
                return redirect(detail_accepted_service, pk)
        except IntegrityError:            
            messages.error(request, "Error removing the accepted warranty application.")            
            return redirect(detail_accepted_service, pk)
    except Http404:
        messages.error(request, "Invalid accepted service.")
        return redirect(list_accepted_services)
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_accepted_service(request, pk):
    try:
        accepted_service = get_object_or_404(AcceptedService, pk = pk)
        accepted_service.delete()
        messages.success(request, 'Successfully deleted accepted service request.')
        return redirect(list_accepted_services)
    except Http404:
        messages.error(request, "Error. Accepted service request not found.")
        return redirect(list_accepted_services)

@user_passes_test(is_superuser, login_url=admin_login)  
@never_cache
def complete_service(request, pk):
    try:
        accepted_service = get_object_or_404(AcceptedService, pk = pk)
        if request.method == "POST":
            remark = request.POST.get("remark")
            file = request.FILES.get("file")
        try:
            completed_service = CompletedService.objects.create(
                id = accepted_service.id,
                application_datetime = accepted_service.application_datetime,
                accepted_datetime = accepted_service.accepted_datetime,
                customer_name = accepted_service.customer_name,
                email = accepted_service.email,                
                address_site = accepted_service.address_site,                
                product = accepted_service.product,
                contact_number = accepted_service.contact_number,
                serial_number = accepted_service.serial_number,
                model_number = accepted_service.model_number,
                service_description = accepted_service.service_description,                
            )
            if remark:
                completed_service.remark = remark
                completed_service.file = file
                completed_service.save()
            if completed_service:                
                accepted_service.delete()
                messages.success(request, "Service completed.")

                subject = "Service Request Completed"
                message = "Service has been completed."            
                html_content = render_to_string('admin_email_templates/service_request_completed.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": completed_service.id,
                    "product": completed_service.product,
                    "serial_number": completed_service.serial_number,
                    "model_number": completed_service.model_number,                
                    "service_description": completed_service.service_description,
                    "customer_name": completed_service.customer_name,
                    "email": completed_service.email,
                    "contact_number": completed_service.contact_number,
                    "alternative_number": completed_service.alternative_number,
                    "address": completed_service.address_site,
                    "remark": completed_service.remark,
                    "file" : completed_service.file
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(completed_service.email)]
                )

                if completed_service.file:

                    # Attach the file from the CompletedService object
                    file_path = completed_service.file.path
                    with open(file_path, 'rb') as file:
                        file_content = file.read()

                    # Determine content type using mimetypes module
                    mime_type, _ = mimetypes.guess_type(file_path)
                    if mime_type is None:
                        mime_type = 'application/octet-stream'

                    # Attach the file with the determined content type
                    email.attach(completed_service.file.name, file_content, mime_type)                        

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(list_accepted_services)
            else:
                messages.error(request, "Service completion failed.")
                return redirect(detail_accepted_service, pk)
            
        except IntegrityError:
            messages.error(request, "This service has already been completed.")
            return redirect(list_completed_services)
    except Http404:
        messages.error(request, "Invalid started service.")
        return redirect(list_accepted_services)


@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_completed_services(request):    
    completed_services = CompletedService.objects.all().order_by("pk")
    context = {
        "completed_services": completed_services,
    }
    return render(request, 'service/list_completed_services.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def completed_services_to_excel(request):
    wb = Workbook()
    ws = wb.active

    completed_services = CompletedService.objects.all().order_by('completion_datetime')

    headers = ["S.No", "Id", "Completion Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number", "Application Datetime", "Acception Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, service in enumerate(completed_services, start=1):
        row = [index, service.id, service.completion_datetime.strftime("%d-%b-%Y %I:%M %p"), service.customer_name, service.contact_number, service.email, service.alternative_number, service.product.name, service.serial_number, service.model_number, service.application_datetime.strftime("%d-%b-%Y %I:%M %p"), service.accepted_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=completed_services.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_completed_service(request, pk):    
    context = {}
    try:
        completed_service = get_object_or_404(CompletedService, pk=pk)
        context.update({"completed_service": completed_service})

        if request.method == "POST":
            remark = request.POST.get("remark")
            if remark != completed_service.remark:
                completed_service.remark = remark
                completed_service.save()
                messages.success(request, "Remark Updated.")

                subject = "Completed Service's Remark Updated."
                message = "Remark of your completed service application has been updated."            
                html_content = render_to_string('admin_email_templates/email.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": completed_service.id,
                    "product": completed_service.product,
                    "serial_number": completed_service.serial_number,
                    "model_number": completed_service.model_number,                
                    "service_description": completed_service.service_description,
                    "customer_name": completed_service.customer_name,
                    "email": completed_service.email,
                    "contact_number": completed_service.contact_number,
                    "alternative_number": completed_service.alternative_number,
                    "address": completed_service.address_site,
                    "remark": completed_service.remark,
                    "file" : completed_service.file
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(completed_service.email)]
                )

                if completed_service.file:

                    # Attach the file from the CompletedService object
                    file_path = completed_service.file.path
                    with open(file_path, 'rb') as file:
                        file_content = file.read()

                    # Determine content type using mimetypes module
                    mime_type, _ = mimetypes.guess_type(file_path)
                    if mime_type is None:
                        mime_type = 'application/octet-stream'

                    # Attach the file with the determined content type
                    email.attach(completed_service.file.name, file_content, mime_type)                        

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(detail_completed_service, pk)

    except Http404:
        messages.error(request, "Invalid. Completed service not found.")
        return redirect(list_completed_services)
    return render(request, 'service/detail_completed_service.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_completed_service(request, pk):
    try:
        completed_service = get_object_or_404(CompletedService, pk = pk)
        completed_service.delete()
        messages.success(request, 'Successfully deleted completed service request.')
        return redirect(list_completed_services)
    except Http404:
        messages.error(request, "Error. Completed service request not found.")
        return redirect(list_completed_services)

@user_passes_test(is_superuser, login_url=admin_login)
@csrf_exempt
@never_cache
def complete_service_file_update(request, pk):
    context = {}
    try:
        completed_service = get_object_or_404(CompletedService, pk=pk)
        context.update({"completed_service": completed_service})

        if request.method == "POST":
            file = request.FILES.get("file")
            completed_service.file = file
            completed_service.save()
            messages.success(request, "File Updated.")

            subject = "Completed Service's File Updated."
            message = "File of your completed service application has been updated."            
            html_content = render_to_string('admin_email_templates/email.html', {
                'subject': subject,
                'message': message,
                "email_subject": subject,
                "email_content": message,
                "id": completed_service.id,
                "product": completed_service.product,
                "serial_number": completed_service.serial_number,
                "model_number": completed_service.model_number,                
                "service_description": completed_service.service_description,
                "customer_name": completed_service.customer_name,
                "email": completed_service.email,
                "contact_number": completed_service.contact_number,
                "alternative_number": completed_service.alternative_number,
                "address": completed_service.address_site,
                "remark": completed_service.remark,
                "file" : completed_service.file
                })            
                    

            email = EmailMessage(
                subject=subject,
                body=html_content,
                from_email="ipcshelpyou@gmail.com",
                to=[str(completed_service.email)]
            )

            if completed_service.file:

                # Attach the file from the CompletedService object
                file_path = completed_service.file.path
                with open(file_path, 'rb') as file:
                    file_content = file.read()

                # Determine content type using mimetypes module
                mime_type, _ = mimetypes.guess_type(file_path)
                if mime_type is None:
                    mime_type = 'application/octet-stream'

                # Attach the file with the determined content type
                email.attach(completed_service.file.name, file_content, mime_type)                        

            # Set the content type of the email to 'text/html'
            email.content_subtype = 'html'

            # Send the email
            email.send()
            
            return redirect(detail_completed_service, pk)
        
    except Http404:
        messages.error(request, "Invalid. Completed service not found.")
        return redirect(list_completed_services)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_completed_service(request, pk):    
    try:
        completed_service = get_object_or_404(CompletedService, pk=pk)
        try:
            accepted_service = AcceptedService.objects.create(
                id = completed_service.id,
                application_datetime = completed_service.application_datetime,
                customer_name = completed_service.customer_name,
                address_site = completed_service.address_site,
                email = completed_service.email,                
                product = completed_service.product,
                contact_number = completed_service.contact_number,
                alternative_number = completed_service.alternative_number,
                serial_number = completed_service.serial_number,
                model_number = completed_service.model_number,
                service_description = completed_service.service_description,
            )
            accepted_service.accepted_datetime = completed_service.accepted_datetime
            accepted_service.save()
            if accepted_service:
                completed_service.delete()
                messages.success(request, "Moved the application back to Accepted Services.")            
                return redirect(list_completed_services)
            else:
                messages.error(request, "Error removing the completed service application.")            
                return redirect(detail_completed_service, pk)
        except IntegrityError:            
            messages.error(request, "Error removing the completed service application.")            
            return redirect(detail_completed_service, pk)
    except Http404:
        messages.error(request, "Invalid completed service.")
        return redirect(list_completed_services)    

# repair
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_repair_requests(request):             
    repair_requests = RepairRequest.objects.all().order_by("-application_datetime")    
    context = {
        "repair_requests": repair_requests,
    }
    return render(request, 'repair/list_repair_requests.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def repair_requests_to_excel(request):
    wb = Workbook()
    ws = wb.active

    repair_requests = RepairRequest.objects.all().order_by('application_datetime')

    headers = ["S.No", "Id", "Application Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, request in enumerate(repair_requests, start=1):
        row = [index, request.id, request.application_datetime.strftime("%d-%b-%Y %I:%M %p"), request.customer_name, request.contact_number, request.email_id, request.alternative_number, request.product.name, request.serial_number, request.model_number]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=repair_requests.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_repair_request(request, pk):
    context = {}
    try:
        repair_request = get_object_or_404(RepairRequest, pk = pk)
        context.update({"repair_request": repair_request})
        try:
            get_object_or_404(StartedRepair, pk = pk)
            context['is_scheduled'] = True
        except Http404:
            pass
    except Http404:
        messages.error(request, "Error. Repair request not found.")        
    return render(request, 'repair/detail_repair_request.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_repair_request(request, pk):
    try:
        repair_request = get_object_or_404(RepairRequest, pk = pk)
        repair_request.delete()
        messages.success(request, 'Successfully deleted repair request.')
        return redirect(list_repair_requests)
    except Http404:
        messages.error(request, "Error. Repair request not found.")
        return redirect(list_repair_requests)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def start_repair(request, pk):    
    # context = {}        
    try:
        get_object_or_404(StartedRepair, pk = pk)
        messages.error(request, "Repair already have started.")
        return redirect(list_started_repairs)
    except:
        pass
    try:
        repair_request = get_object_or_404(RepairRequest, pk = pk)      
        try:
            started_repair = StartedRepair.objects.create(
                id = repair_request.id,
                application_datetime = repair_request.application_datetime,
                customer_name = repair_request.customer_name,
                address_customer = repair_request.address_customer,
                email_id = repair_request.email_id,                
                product = repair_request.product,
                contact_number = repair_request.contact_number,
                alternative_number = repair_request.alternative_number,
                serial_number = repair_request.serial_number,
                model_number = repair_request.model_number,
                item_description = repair_request.item_description,
            ) 
            if started_repair:
                repair_request.delete()                       
                messages.success(request, "Repair started!")
                
                subject = "Repair Request Started"
                message = "Repair has been started for your repair request."
                html_content = render_to_string('admin_email_templates/repair_request_started.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": started_repair.id,
                    "product": started_repair.product,
                    "serial_number": started_repair.serial_number,
                    "model_number": started_repair.model_number,                                    
                    "repair_description": started_repair.item_description,
                    "customer_name": started_repair.customer_name,
                    "email": started_repair.email_id,
                    "contact_number": started_repair.contact_number,
                    "alternative_number": started_repair.alternative_number,
                    "address": started_repair.address_customer,                    
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(started_repair.email_id)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(list_repair_requests)            
        except IntegrityError:                
            messages.warning(request, "Repair has already started for this repair request.")
            return redirect(detail_started_repair, pk)                
    except Http404:
        messages.error(request, "Invalid repair request.")
        return redirect(list_repair_requests)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_started_repair(request, pk):    
    try:
        started_repair = get_object_or_404(StartedRepair, pk=pk)
        try:
            repair_request = RepairRequest.objects.create(
                id = started_repair.id,
                customer_name = started_repair.customer_name,
                address_customer = started_repair.address_customer,
                email_id = started_repair.email_id,
                product = started_repair.product,                
                contact_number = started_repair.contact_number,
                alternative_number = started_repair.alternative_number,
                serial_number = started_repair.serial_number,
                model_number = started_repair.model_number,
                item_description = started_repair.item_description
                )            
            repair_request.application_datetime = started_repair.application_datetime
            repair_request.save()
            if repair_request:
                started_repair.delete()
                messages.success(request, "Application has been moved back to Repair Request stage.")
                return redirect(list_started_repairs)
            else:
                messages.error(request, "Error removing started repair application.")
                return redirect(detail_started_repair,  pk)   
        except IntegrityError:
            messages.error(request, "Error removing started repair application.")
            return redirect(detail_started_repair,  pk)                
    except Http404:
        messages.error(request, "Invalid started repair.")
        return redirect(list_started_repairs)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_started_repairs(request):    
    started_repairs = StartedRepair.objects.all().order_by("-started_datetime")    
    context = {        
        "started_repairs": started_repairs
    }
    return render(request, 'repair/list_started_repairs.html', context)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def started_repairs_to_excel(request):
    wb = Workbook()
    ws = wb.active

    started_repairs = StartedRepair.objects.all().order_by('started_datetime')

    headers = ["S.No", "Id", "Started Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number", "Application Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, repair in enumerate(started_repairs, start=1):
        row = [index, repair.id, repair.started_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.customer_name, repair.contact_number, repair.email_id, repair.alternative_number, repair.product.name, repair.serial_number, repair.model_number, repair.application_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=started_repairs.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_started_repair(request, pk):    
    try:
        started_repair = get_object_or_404(StartedRepair, pk = pk)
        context = {"started_repair" : started_repair}
    except Http404:
        messages.error(request, "Invalid started repair.")
        return redirect(list_started_repairs)
    return render(request, 'repair/detail_started_repair.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_started_repair(request, pk):
    try:
        started_repair = get_object_or_404(StartedRepair, pk = pk)
        started_repair.delete()
        messages.success(request, 'Successfully deleted started repair request.')
        return redirect(list_started_repairs)
    except Http404:
        messages.error(request, "Error. Started repair request not found.")
        return redirect(list_started_repairs)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def complete_repair(request, pk):    
    try:
        started_repair = get_object_or_404(StartedRepair, pk = pk)
        try:
            completed_repair = CompletedRepair.objects.create(
                id = started_repair.id,
                application_datetime = started_repair.application_datetime,
                started_datetime = started_repair.started_datetime,
                customer_name = started_repair.customer_name,
                email_id = started_repair.email_id,                
                alternative_number = started_repair.alternative_number,
                address_customer = started_repair.address_customer,
                product = started_repair.product,
                contact_number = started_repair.contact_number,
                serial_number = started_repair.serial_number,
                model_number = started_repair.model_number,
                item_description = started_repair.item_description,                
            )
            try:
                get_object_or_404(CompletedRepair, pk = pk)
                started_repair.delete()
                messages.success(request, "Repair completed.")

                subject = "Repair Request Completed"
                message = "The repair has been completed."
                html_content = render_to_string('admin_email_templates/repair_request_completed.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": completed_repair.id,
                    "product": completed_repair.product,
                    "serial_number": completed_repair.serial_number,
                    "model_number": completed_repair.model_number,                                    
                    "repair_description": completed_repair.item_description,
                    "customer_name": completed_repair.customer_name,
                    "email": completed_repair.email_id,
                    "contact_number": completed_repair.contact_number,
                    "alternative_number": completed_repair.alternative_number,
                    "address": completed_repair.address_customer,                    
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(started_repair.email_id)]
                )                    

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(list_started_repairs)        
            except Http404:
                messages.error(request, "Repair completion failed.")
                return redirect(detail_started_repair, pk)
            
        except IntegrityError:
            messages.error(request, "This repair has already been completed.")
            return redirect(list_completed_repairs)
    except Http404:
        messages.error(request, "Invalid started repair.")
        return redirect(list_started_repairs)    

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_completed_repairs(request):    
    completed_repairs = CompletedRepair.objects.all().order_by("pk")    
    context = {
        "completed_repairs": completed_repairs,
    }
    return render(request, 'repair/list_completed_repairs.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def completed_repairs_to_excel(request):
    wb = Workbook()
    ws = wb.active

    completed_repairs = CompletedRepair.objects.all().order_by('completion_datetime')

    headers = ["S.No", "Id", "Completion Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number", "Application Datetime", "Started Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, repair in enumerate(completed_repairs, start=1):
        row = [index, repair.id, repair.completion_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.customer_name, repair.contact_number, repair.email_id, repair.alternative_number, repair.product.name, repair.serial_number, repair.model_number, repair.application_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.started_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=completed_repairs.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_completed_repair(request, pk):    
    context = {}
    try:
        completed_repair = get_object_or_404(CompletedRepair, pk=pk)
        context.update({
            "completed_repair": completed_repair
        })
    except Http404:
        messages.error(request, "Invalid. Completed repair not found.")
        return redirect(list_completed_repairs)
    return render(request, 'repair/detail_completed_repair.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_completed_repair(request, pk):    
    try:
        completed_repair = get_object_or_404(CompletedRepair, pk=pk)
        try:
            started_repair = StartedRepair.objects.create(
                id = completed_repair.id,
                application_datetime = completed_repair.application_datetime,
                customer_name = completed_repair.customer_name,
                address_customer = completed_repair.address_customer,
                email_id = completed_repair.email_id,
                product = completed_repair.product,                
                contact_number = completed_repair.contact_number,
                alternative_number = completed_repair.alternative_number,
                serial_number = completed_repair.serial_number,
                model_number = completed_repair.model_number,
                item_description = completed_repair.item_description
                )            
            started_repair.started_datetime = completed_repair.started_datetime
            started_repair.save()
            if started_repair:
                completed_repair.delete()
                messages.success(request, "Application has been moved back to Started Repair stage.")
                return redirect(list_completed_repairs)
            else:
                messages.error(request, "Error removing completed repair application.")
                return redirect(detail_completed_repair,  pk)   
        except IntegrityError:
            messages.error(request, "Error removing completed repair application.")
            return redirect(detail_completed_repair,  pk)                 
    except Http404:
        messages.error(request, "Invalid completed repair.")
        return redirect(list_completed_repairs)
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_completed_repair(request, pk):
    try:
        completed_repair = get_object_or_404(CompletedRepair, pk = pk)
        completed_repair.delete()
        messages.success(request, 'Successfully deleted completed repair request.')
        return redirect(list_completed_repairs)
    except Http404:
        messages.error(request, "Error. Completed repair request not found.")
        return redirect(list_completed_repairs)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def set_repair_ready_to_dispatch(request, pk):    
    try:
        completed_repair = get_object_or_404(CompletedRepair, pk = pk)
        if request.method == "POST":
            remark = request.POST.get('remark')
            file = request.FILES.get('file')
            try:
                dispatch_ready_item = RepairReadyToDispatch.objects.create(
                    id = completed_repair.id,
                    application_datetime = completed_repair.application_datetime,
                    started_datetime = completed_repair.started_datetime,
                    completion_datetime = completed_repair.completion_datetime,
                    customer_name = completed_repair.customer_name,
                    email_id = completed_repair.email_id,                
                    alternative_number = completed_repair.alternative_number,
                    address_customer = completed_repair.address_customer,
                    product = completed_repair.product,
                    contact_number = completed_repair.contact_number,
                    serial_number = completed_repair.serial_number,
                    model_number = completed_repair.model_number,
                    item_description = completed_repair.item_description,
                    remark = remark,
                    file = file
                )
                try:
                    get_object_or_404(CompletedRepair, pk = pk)                
                    completed_repair.delete()                
                    messages.success(request, "Repaired item is ready to dispatch.")
                    
                    subject = "Repair Request Ready to Dispatch"
                    message = "The repair item is ready for dispatching."
                    html_content = render_to_string('admin_email_templates/repair_request_ready_to_dispatch.html', {
                        'subject': subject,
                        'message': message,
                        "email_subject": subject,
                        "email_content": message,
                        "id": dispatch_ready_item.id,
                        "product": dispatch_ready_item.product,
                        "serial_number": dispatch_ready_item.serial_number,
                        "model_number": dispatch_ready_item.model_number,                
                        "repair_description": dispatch_ready_item.item_description,
                        "customer_name": dispatch_ready_item.customer_name,
                        "email": dispatch_ready_item.email_id,
                        "contact_number": dispatch_ready_item.contact_number,
                        "alternative_number": dispatch_ready_item.alternative_number,
                        "address": dispatch_ready_item.address_customer,
                        "remark": dispatch_ready_item.remark,
                        "file" : dispatch_ready_item.file
                        })            
                            

                    email = EmailMessage(
                        subject=subject,
                        body=html_content,
                        from_email="ipcshelpyou@gmail.com",
                        to=[str(dispatch_ready_item.email_id)]
                    )

                    if dispatch_ready_item.file:

                        # Attach the file from the CompletedService object
                        file_path = dispatch_ready_item.file.path
                        with open(file_path, 'rb') as file:
                            file_content = file.read()

                        # Determine content type using mimetypes module
                        mime_type, _ = mimetypes.guess_type(file_path)
                        if mime_type is None:
                            mime_type = 'application/octet-stream'

                        # Attach the file with the determined content type
                        email.attach(dispatch_ready_item.file.name, file_content, mime_type)                        

                    # Set the content type of the email to 'text/html'
                    email.content_subtype = 'html'

                    # Send the email
                    email.send()

                    return redirect(list_completed_repairs)        
                except Http404:
                    messages.error(request, "Failed to set as ready to dispatch.")
                    return redirect(detail_completed_repair, pk)
                
            except IntegrityError:
                messages.error(request, "This repair item has already been set as ready to dispatch.")
                return redirect(list_completed_repairs)
    except Http404:
        messages.error(request, "Invalid completed repair.")
        return redirect(list_completed_repairs)
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def list_repairs_ready_to_dispatch(request):    
    dispatch_ready_items = RepairReadyToDispatch.objects.all().order_by("-dispatch_ready_datetime")
    context = {        
        "dispatch_ready_items": dispatch_ready_items
    }
    return render(request, 'repair/list_repair_ready_to_dispatches.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def repairs_ready_to_dispatch_to_excel(request):
    wb = Workbook()
    ws = wb.active

    repairs_ready_to_dispatch = RepairReadyToDispatch.objects.all().order_by('completion_datetime')

    headers = ["S.No", "Id", "Ready For Dispatch Datetime", "Customer Name", "Contact Number", "Email ID", "Alternative Number", "Product", "Serial Number", "Model Number", "Application Datetime", "Started Datetime", "Completion Datetime"]
    ws.append(headers)
    ws.append([])

    for cell in ws[1]:  # Iterate over cells in the first row (headers)
        cell.alignment = Alignment(horizontal='center')

    # Append data rows
    for index, repair in enumerate(repairs_ready_to_dispatch, start=1):
        row = [index, repair.id, repair.dispatch_ready_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.customer_name, repair.contact_number, repair.email_id, repair.alternative_number, repair.product.name, repair.serial_number, repair.model_number, repair.application_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.started_datetime.strftime("%d-%b-%Y %I:%M %p"), repair.completion_datetime.strftime("%d-%b-%Y %I:%M %p")]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Automatically adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Set response for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=repairs_ready_to_dispatch.xlsx'
    wb.save(response)

    return response

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def detail_repair_ready_to_dispatch(request, pk):    
    context = {}
    try:
        dispatch_ready_item = get_object_or_404(RepairReadyToDispatch, pk=pk)
        context.update({"dispatch_ready_item": dispatch_ready_item})

        if request.method == "POST":
            remark = request.POST.get("remark")
            if remark != dispatch_ready_item.remark:
                dispatch_ready_item.remark = remark
                dispatch_ready_item.save()
                messages.success(request, "Remark Updated.")            

                subject = "Repair Item Ready To Dispatch's Remark Updated."
                message = "The remark of repair item which is ready to dispatch has been updated."
                html_content = render_to_string('admin_email_templates/email.html', {
                    'subject': subject,
                    'message': message,
                    "email_subject": subject,
                    "email_content": message,
                    "id": dispatch_ready_item.id,
                    "product": dispatch_ready_item.product,
                    "serial_number": dispatch_ready_item.serial_number,
                    "model_number": dispatch_ready_item.model_number,                
                    "item_description": dispatch_ready_item.item_description,
                    "customer_name": dispatch_ready_item.customer_name,
                    "email": dispatch_ready_item.email_id,
                    "contact_number": dispatch_ready_item.contact_number,
                    "alternative_number": dispatch_ready_item.alternative_number,
                    "address": dispatch_ready_item.address_customer,
                    "remark": dispatch_ready_item.remark,
                    "file" : dispatch_ready_item.file
                    })            
                        

                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email="ipcshelpyou@gmail.com",
                    to=[str(dispatch_ready_item.email_id)]
                )

                if dispatch_ready_item.file:

                    # Attach the file from the CompletedService object
                    file_path = dispatch_ready_item.file.path
                    with open(file_path, 'rb') as file:
                        file_content = file.read()

                    # Determine content type using mimetypes module
                    mime_type, _ = mimetypes.guess_type(file_path)
                    if mime_type is None:
                        mime_type = 'application/octet-stream'

                    # Attach the file with the determined content type
                    email.attach(dispatch_ready_item.file.name, file_content, mime_type)                        

                # Set the content type of the email to 'text/html'
                email.content_subtype = 'html'

                # Send the email
                email.send()

                return redirect(detail_repair_ready_to_dispatch, pk)

    except Http404:
        messages.error(request, "Invalid. Repair item that has been ready to dispatch is not found.")
        return redirect(list_repairs_ready_to_dispatch)
    return render(request, 'repair/detail_repair_ready_to_dispatch.html', context)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def update_repair_ready_to_dispatch_file(request, pk):
    context = {}
    try:
        dispatch_ready_item = get_object_or_404(RepairReadyToDispatch, pk=pk)
        context.update({"dispatch_ready_item": dispatch_ready_item})

        if request.method == "POST":
            file = request.FILES.get("file")
            dispatch_ready_item.file = file
            dispatch_ready_item.save()
            messages.success(request, "File Updated.")

            # subject = "File of repair ready to dispatch has been updated."
            # message = "File has been updated."
            # html_content = render_to_string('email_templates/file_updation_completed_service.html', {'subject': subject, 'message': message, 'completed_service': completed_service})

            # email = EmailMessage(
            #     subject=subject,
            #     body=html_content,
            #     from_email="ipcshelpyou@gmail.com",
            #     to=[str(completed_service.email)]
            # )

            # # Attach the file from the CompletedService object
            # file_path = completed_service.file.path
            # with open(file_path, 'rb') as file:
            #     file_content = file.read()

            # # Determine content type using mimetypes module
            # mime_type, _ = mimetypes.guess_type(file_path)
            # if mime_type is None:
            #     mime_type = 'application/octet-stream'

            # # Attach the file with the determined content type
            # email.attach(completed_service.file.name, file_content, mime_type)

            # # Set the content type of the email to 'text/html'
            # email.content_subtype = 'html'

            # # Send the email
            # email.send()

            subject = "Repair Item Ready To Dispatch's File Updated."
            message = "The file of repair item which is ready to dispatch has been updated."
            html_content = render_to_string('admin_email_templates/email.html', {
                'subject': subject,
                'message': message,
                "email_subject": subject,
                "email_content": message,
                "id": dispatch_ready_item.id,
                "product": dispatch_ready_item.product,
                "serial_number": dispatch_ready_item.serial_number,
                "model_number": dispatch_ready_item.model_number,                
                "item_description": dispatch_ready_item.item_description,
                "customer_name": dispatch_ready_item.customer_name,
                "email": dispatch_ready_item.email_id,
                "contact_number": dispatch_ready_item.contact_number,
                "alternative_number": dispatch_ready_item.alternative_number,
                "address": dispatch_ready_item.address_customer,
                "remark": dispatch_ready_item.remark,
                "file" : dispatch_ready_item.file
                })            
                    

            email = EmailMessage(
                subject=subject,
                body=html_content,
                from_email="ipcshelpyou@gmail.com",
                to=[str(dispatch_ready_item.email_id)]
            )

            if dispatch_ready_item.file:

                # Attach the file from the CompletedService object
                file_path = dispatch_ready_item.file.path
                with open(file_path, 'rb') as file:
                    file_content = file.read()

                # Determine content type using mimetypes module
                mime_type, _ = mimetypes.guess_type(file_path)
                if mime_type is None:
                    mime_type = 'application/octet-stream'

                # Attach the file with the determined content type
                email.attach(dispatch_ready_item.file.name, file_content, mime_type)                        

            # Set the content type of the email to 'text/html'
            email.content_subtype = 'html'

            # Send the email
            email.send()

            return redirect(detail_repair_ready_to_dispatch, pk)
    
    except Http404:
        messages.error(request, "Invalid. Repair item ready to dispatch not found.")
        return redirect(list_repairs_ready_to_dispatch)

@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def delete_repair_ready_to_dispatch(request, pk):
    try:
        repair_ready_to_dispatch = get_object_or_404(RepairReadyToDispatch, pk=pk)
        try:
            completed_repair = CompletedRepair.objects.create(
                id = repair_ready_to_dispatch.id,
                application_datetime = repair_ready_to_dispatch.application_datetime,
                started_datetime = repair_ready_to_dispatch.started_datetime,
                customer_name = repair_ready_to_dispatch.customer_name,
                address_customer = repair_ready_to_dispatch.address_customer,
                email_id = repair_ready_to_dispatch.email_id,
                product = repair_ready_to_dispatch.product,                
                contact_number = repair_ready_to_dispatch.contact_number,
                alternative_number = repair_ready_to_dispatch.alternative_number,
                serial_number = repair_ready_to_dispatch.serial_number,
                model_number = repair_ready_to_dispatch.model_number,
                item_description = repair_ready_to_dispatch.item_description
                )                        
            completed_repair.completion_datetime = repair_ready_to_dispatch.completion_datetime
            completed_repair.save()
            if completed_repair:
                repair_ready_to_dispatch.delete()
                messages.success(request, "Application has been moved back to Completed Repair stage.")
                return redirect(list_repairs_ready_to_dispatch)
            else:
                messages.error(request, "Error removing repair ready to dispatch.")
                return redirect(detail_repair_ready_to_dispatch,  pk)   
        except IntegrityError:
            messages.error(request, "Error removing repair ready to dispatch.")
            return redirect(detail_repair_ready_to_dispatch,  pk)                 
    except Http404:
        messages.error(request, "Invalid repair ready to dispatch.")
        return redirect(list_repairs_ready_to_dispatch)    
    
@user_passes_test(is_superuser, login_url=admin_login)
@never_cache
def hard_delete_repair_ready_to_dispatch(request, pk):
    try:
        repair_ready_to_dispatch = get_object_or_404(RepairReadyToDispatch, pk = pk)
        repair_ready_to_dispatch.delete()
        messages.success(request, 'Successfully deleted dispatch ready repair request.')
        return redirect(list_repairs_ready_to_dispatch)
    except Http404:
        messages.error(request, "Error. Repair ready to dispatch not found.")
        return redirect(list_repairs_ready_to_dispatch)

@never_cache
def mark_as_viewed_service(request, pk):
    try:
        service_request = get_object_or_404(ServiceRequest, pk=pk)
        service_request.viewed = True
        service_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass

@never_cache
def mark_as_unread_service(request, pk):
    try:
        service_request = get_object_or_404(ServiceRequest, pk=pk)
        service_request.viewed = False
        service_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass


@never_cache
def mark_as_viewed_repair(request, pk):
    try:
        repair_request = get_object_or_404(RepairRequest, pk=pk)
        repair_request.viewed = True
        repair_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass

@never_cache
def mark_as_unread_repair(request, pk):
    try:
        repair_request = get_object_or_404(RepairRequest, pk=pk)
        repair_request.viewed = False
        repair_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass


@never_cache
def mark_as_viewed_warranty(request, pk):
    try:
        warranty_request = get_object_or_404(WarrantyApplication, pk=pk)
        warranty_request.viewed = True
        warranty_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass

@never_cache
def mark_as_unread_warranty(request, pk):
    try:
        warranty_request = get_object_or_404(WarrantyApplication, pk=pk)
        warranty_request.viewed = False
        warranty_request.save()
        return JsonResponse({"message": 'success'})
    except Http404:
        pass
